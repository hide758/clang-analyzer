from django.core.management.base import BaseCommand
from ...models import Project, Function, FunctionRelation

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import logging
import datetime
import pathlib
import copy

logger = logging.getLogger('Survey')


class Command(BaseCommand):
    """exportdb command class

    Args:
        BaseCommand (_type_): Django base command class
    """    
    help = "Export Database"
    _template = "template_exportdb.xlsx"

    def _MakeFunctionList(self, ProjectName:str, ws:openpyxl.worksheet.worksheet.Worksheet):
        """Make function list

        Args:
            ProjectName (str): Project name
            ws (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet

        Returns:
            _type_: Registered function count
        """        
        Functions = Function.objects.filter(project__name=ProjectName)

        if Functions.count() == 0:
            return 0
        
        # make summary
        summary = {
            "{作成日時}"      : datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S'),
            "{解析完了日時}"  : Functions.order_by("modified").first().modified.astimezone().strftime('%Y/%m/%d %H:%M:%S'),
            "{プロジェクト}"  : ProjectName,
            "{関数数}"        : Functions.count(),
            }


        # make function list
        no = 1
        func_list = [["No.", "プロジェクト", "const", "static", "戻り値型", "関数名", "引数", "ファイル", "先頭行番号", "末尾行番号"],]
        for obj in Functions:
            args = "\n".join([f"{ag['Type']} {ag['Name']}" for ag in obj.arguments])
            func_list.append([
                no,                             # No.
                obj.project.name,               # プロジェクト
                "✓" if obj.const else "",      # (呼び出し元) const
                "✓" if obj.static else "",     # (呼び出し元) const
                obj.return_type,                # 戻り値型
                obj.name,                       # 関数名
                args,                           # 引数
                obj.file,                       # ファイル
                obj.line,                       # 先頭行番号
                obj.end_line,                   # 末尾行番号
                ])
            
            no += 1

        # scan template and set data
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # search and replace summary
                if cell.value is not None and cell.value in summary.keys():
                    cell.value = summary[cell.value]
                
                # get base row of function list
                if cell.value is not None and cell.value == "{関数一覧}":
                    func_list_row = cell.row
                    func_list_col = cell.column

        # write header
        for index, header in enumerate(func_list[0]):
            ws.cell(row=func_list_row, column=func_list_col + index).value = header

        # write function list
        for row_index, row in enumerate(func_list[1:], start=1):
            for col_index, col in enumerate(row):
                # get cells
                currentcell = ws.cell(row=func_list_row + row_index, column=func_list_col + col_index)
                belowcell = ws.cell(row=func_list_row + row_index + 1, column=func_list_col + col_index)

                # copy format to below cell
                belowcell.font = copy.copy(currentcell.font)
                belowcell.number_format = copy.copy(currentcell.number_format)
                belowcell.alignment = copy.copy(currentcell.alignment)        

                # set item
                currentcell.value = col

        # get all records
        func_table = Table(displayName='関数一覧',
                           ref=f"{get_column_letter(func_list_col)}{func_list_row}:{get_column_letter(func_list_col + len(func_list[0]) - 1)}{func_list_row + len(func_list) - 1}")
        func_table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium8', showRowStripes=True)
        ws.add_table(func_table)

        return Functions.count()


    def _MakeFunctionRelationList(self, ProjectName:str, ws:openpyxl.worksheet.worksheet.Worksheet):
        """Make function relation list

        Args:
            ProjectName (str): Project name
            ws (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet

        Returns:
            _type_: Registered function relation count.
        """        
        FunctionRelations = FunctionRelation.objects.filter(project__name=ProjectName)

        if FunctionRelations.count() == 0:
            return 0
        
        # make summary
        summary = {
            "{作成日時}"      : datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S'),
            "{解析完了日時}"  : FunctionRelations.order_by("modified").first().modified.astimezone().strftime('%Y/%m/%d %H:%M:%S'),
            "{プロジェクト}"  : ProjectName,
            "{呼び出し数}"      : FunctionRelations.count(),
            }


        # make function list
        no = 1
        func_list = [["No.", "プロジェクト", "const\n(呼び出し元)", "static\n(呼び出し元)", "戻り値\n(呼び出し元)", "関数名\n(呼び出し元)", "const\n(呼び出し先)", "static\n(呼び出し先)",  "戻り値\n(呼び出し先)","関数名\n(呼び出し先)", "ファイル", "行番号"],]
        for obj in FunctionRelations:
            func_list.append([
                no,                                         # No.
                obj.project.name,                           # プロジェクト
                "✓" if obj.call_from.const else "",     # (呼び出し元) const
                "✓" if obj.call_from.static else "",    # (呼び出し元) static
                obj.call_from.return_type,               # (呼び出し元) 戻り値型
                obj.call_from.name,                      # (呼び出し元) 関数名
                "✓" if obj.call_to.const else "",       # (呼び出し先) const
                "✓" if obj.call_to.static else "",      # (呼び出し先) static
                obj.call_to.return_type,                # (呼び出し先) 戻り値型
                obj.call_to.name,                        # (呼び出し先) 関数名
                obj.file,                                   # ファイル
                obj.line,                                   # 行番号
                ])
            
            no += 1

        # scan template and set summary data
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # search and replace summary
                if cell.value is not None and cell.value in summary.keys():
                    cell.value = summary[cell.value]
                
                # get base row of function list
                if cell.value is not None and cell.value == "{呼び出し一覧}":
                    func_list_row = cell.row
                    func_list_col = cell.column

        # write header
        for index, header in enumerate(func_list[0]):
            ws.cell(row=func_list_row, column=func_list_col + index).value = header

        # write function relation list
        for row_index, row in enumerate(func_list[1:], start=1):
            for col_index, col in enumerate(row):
                # get cells
                currentcell = ws.cell(row=func_list_row + row_index, column=func_list_col + col_index)
                belowcell = ws.cell(row=func_list_row + row_index + 1, column=func_list_col + col_index)

                # copy format to below cell
                belowcell.font = copy.copy(currentcell.font)
                belowcell.number_format = copy.copy(currentcell.number_format)
                belowcell.alignment = copy.copy(currentcell.alignment)        

                # set item
                currentcell.value = col

        # get all records
        func_table = Table(displayName='呼び出し一覧',
                           ref=f"{get_column_letter(func_list_col)}{func_list_row}:{get_column_letter(func_list_col + len(func_list[0]) - 1)}{func_list_row + len(func_list) - 1}")
        func_table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium8', showRowStripes=True)
        ws.add_table(func_table)


        return FunctionRelations.count()



    def _ToExcel(self, ProjectName:str, SaveAs:str):
        """export to Excel file

        Args:
            ProjectName (str): Project Name
            SaveAs (str): Save file
        """        
        # open template file
        wb = load_workbook(pathlib.Path(__file__).resolve().parent / self._template)


        # export functions
        ws = wb["関数一覧"]
        funccnt = self._MakeFunctionList(ProjectName, ws)

        if funccnt == 0:
            logger.info(" No function exported")

        else:
            logger.info(f" {funccnt} functions exported")

        # export function relations
        ws = wb["呼び出し一覧"]
        funcrelcnt = self._MakeFunctionRelationList(ProjectName, ws)
        if funccnt == 0:
            logger.info(" No function relations exported")

        else:
            logger.info(f" {funcrelcnt} function relations exported")

        
        wb.save(filename=SaveAs)

    def handle(self, *args, **options):
        """command entry point

        """        
        try:
            start_time = datetime.datetime.now()

            logger.info("Export database Start")
            logger.info(f" {start_time.strftime('%Y/%m/%d %H:%M:%S')}")
            

            self._ToExcel(ProjectName=options['project'], SaveAs=options['save_as'])
            
        except Exception as e:
            logger.error(f"exception {e}", exc_info=True)


        finally:
            logger.debug(f" elapsed time {(datetime.datetime.now() - start_time).total_seconds()}")

    def add_arguments(self, parser):
        """regist command arguments

        Args:
            parser (_type_): argument parser
        """        

        parser.add_argument('--project', nargs='?', default=None, type=str)
        parser.add_argument('--format', nargs='?', default="Excel", type=str)
        parser.add_argument('--save-as', nargs='?', default="clangAnalyzer.xlsx", type=str)
